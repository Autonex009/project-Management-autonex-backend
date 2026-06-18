"""
Onboarding API - Endpoints for modules, sections, quizzes, candidate progress, and team dashboard.
"""
import csv
import io
import json
from datetime import datetime
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Response, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel

from app.db.database import get_db
from app.models.user import User
from app.models.employee import Employee
from app.models.onboarding import (
    OnboardingModule,
    OnboardingSection,
    OnboardingDocument,
    OnboardingQuizQuestion,
    OnboardingProgress,
    OnboardingQuizAttempt,
    OnboardingTeamMember,
)
from app.schemas.onboarding import (
    OnboardingModuleCreate,
    OnboardingModuleResponse,
    OnboardingSectionCreate,
    OnboardingSectionResponse,
    OnboardingQuizAttemptCreate,
    OnboardingQuizAttemptResponse,
    OnboardingProgressCreate,
    OnboardingProgressResponse,
    OnboardingTeamMemberCreate,
    OnboardingTeamMemberResponse,
    OnboardingDocumentCreate,
    OnboardingQuizQuestionCreate,
)
from app.services.auth_service import get_current_user, require_role

router = APIRouter(prefix="/api/onboarding", tags=["Onboarding"])

# Try to import openpyxl for Excel imports; raise HTTP error if missing when route is called
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ── Helper: Get Candidates ──────────────────────────────────────────
def get_onboarding_candidates(db: Session) -> List[User]:
    """Retrieve all users whose designation is Annotator or Reviewer."""
    return db.query(User).join(
        Employee, User.employee_id == Employee.id
    ).filter(
        (Employee.designation.ilike("%Annotator%")) | (Employee.designation.ilike("%Reviewer%"))
    ).all()


def is_module_locked(user_id: int, module_id: int, db: Session) -> bool:
    """Check if a module is locked for a candidate based on sequence completion."""
    # Retrieve all published modules sorted by order
    modules = db.query(OnboardingModule).filter(OnboardingModule.status.ilike("PUBLISHED")).order_by(OnboardingModule.order.asc()).all()
    
    previous_completed = True
    for m in modules:
        if m.id == module_id:
            return not previous_completed
            
        total_sections = len(m.sections)
        completed_count = db.query(OnboardingProgress).filter(
            OnboardingProgress.user_id == user_id,
            OnboardingProgress.module_id == m.id
        ).count()
        
        is_completed = (total_sections > 0 and completed_count == total_sections) or (total_sections == 0)
        previous_completed = is_completed
        
    return False


# ── Serialization helpers (control quiz-answer exposure) ─────────────
# The correct answer (correct_option_index) must never reach the candidate's
# browser, and is not needed by any list view. It is included only when an
# admin/PM fetches a single module to edit it in the builder.

def _serialize_question(question: OnboardingQuizQuestion, include_answer: bool) -> dict:
    data = {
        "id": question.id,
        "section_id": question.section_id,
        "question": question.question,
        "options": question.options,
    }
    if include_answer:
        data["correct_option_index"] = question.correct_option_index
    return data


def _serialize_section(section: OnboardingSection, include_answers: bool) -> dict:
    return {
        "id": section.id,
        "module_id": section.module_id,
        "title": section.title,
        "description": section.description,
        "video_url": section.video_url,
        "video_duration": section.video_duration,
        "quiz_passing_score": section.quiz_passing_score,
        "order": section.order,
        "documents": [
            {"id": d.id, "section_id": d.section_id, "title": d.title, "type": d.type, "url": d.url}
            for d in section.documents
        ],
        "questions": [_serialize_question(q, include_answers) for q in section.questions],
    }


def _serialize_module(module: OnboardingModule, include_answers: bool) -> dict:
    return {
        "id": module.id,
        "title": module.title,
        "description": module.description,
        "status": module.status,
        "assessment_url": module.assessment_url,
        "order": module.order,
        "created_at": module.created_at,
        "updated_at": module.updated_at,
        "sections": [
            _serialize_section(s, include_answers)
            for s in sorted(module.sections, key=lambda x: (x.order or 0))
        ],
    }


# ── Modules Endpoints ───────────────────────────────────────────────

@router.get("/modules")
def get_modules(
    include_drafts: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all onboarding modules sorted by order. Quiz answers are never included in the list."""
    query = db.query(OnboardingModule)
    if not include_drafts and current_user.role == "employee":
        query = query.filter(OnboardingModule.status.ilike("PUBLISHED"))
    modules = query.order_by(OnboardingModule.order.asc()).all()
    return [_serialize_module(m, include_answers=False) for m in modules]


# ── Excel Import / Templates (placed above to avoid shadowing) ──────────

@router.get("/modules/quiz-sample-excel")
def download_quiz_sample():
    """Generate and return a sample Excel template for quiz import."""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=501, detail="Excel library (openpyxl) is not installed on the server.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quiz Questions"

    # Header
    headers = ['question', 'option1', 'option2', 'option3', 'option4', 'correctOption']
    ws.append(headers)

    # Samples
    ws.append(['What year was the company founded?', '2005', '2008', '2010', '2012', '2'])
    ws.append(['What is our core value?', 'Speed', 'Innovation', 'Profit', 'Scale', '2'])
    ws.append(['Who is the CEO?', 'John', 'Jane', 'Mike', 'Sarah', '1'])

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=autonex_quiz_sample.xlsx"}
    )


@router.get("/modules/{module_id}")
def get_module(
    module_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Retrieve a single module with its sections, documents, and questions.

    Correct quiz answers are returned only to admins/PMs (the module builder needs
    them to edit); candidates never receive them.
    """
    module = db.query(OnboardingModule).filter(OnboardingModule.id == module_id).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")

    if current_user.role == "employee":
        if is_module_locked(current_user.id, module_id, db):
            raise HTTPException(status_code=403, detail="This module is locked.")

    include_answers = current_user.role in ("admin", "pm")
    return _serialize_module(module, include_answers=include_answers)


@router.post("/modules", response_model=OnboardingModuleResponse, status_code=status.HTTP_201_CREATED)
def create_module(
    payload: OnboardingModuleCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Create a new module. Restricted to Admins and PMs."""
    module_data = payload.model_dump(exclude={"sections"})
    module = OnboardingModule(**module_data)
    db.add(module)
    db.flush()

    if payload.sections:
        for idx, sec in enumerate(payload.sections):
            section = OnboardingSection(
                module_id=module.id,
                title=sec.title,
                description=sec.description,
                video_url=sec.video_url,
                video_duration=sec.video_duration,
                quiz_passing_score=sec.quiz_passing_score,
                order=sec.order or idx,
            )
            db.add(section)
            db.flush()

            if sec.documents:
                for doc in sec.documents:
                    db_doc = OnboardingDocument(section_id=section.id, **doc.model_dump())
                    db.add(db_doc)

            if sec.questions:
                for q in sec.questions:
                    db_q = OnboardingQuizQuestion(
                        section_id=section.id,
                        question=q.question,
                        options=q.options,
                        correct_option_index=q.correct_option_index
                    )
                    db.add(db_q)

    db.commit()
    db.refresh(module)
    return module


@router.put("/modules/{module_id}", response_model=OnboardingModuleResponse)
def update_module(
    module_id: int,
    payload: OnboardingModuleCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Update a module and replace its nested sections/docs/questions. Restricted to Admins and PMs."""
    module = db.query(OnboardingModule).filter(OnboardingModule.id == module_id).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")

    # Update metadata
    module.title = payload.title
    module.description = payload.description
    module.status = payload.status
    module.assessment_url = payload.assessment_url
    module.order = payload.order

    # Delete existing sections (causes cascade delete of documents and questions)
    db.query(OnboardingSection).filter(OnboardingSection.module_id == module_id).delete(synchronize_session=False)
    db.flush()

    if payload.sections:
        for idx, sec in enumerate(payload.sections):
            section = OnboardingSection(
                module_id=module.id,
                title=sec.title,
                description=sec.description,
                video_url=sec.video_url,
                video_duration=sec.video_duration,
                quiz_passing_score=sec.quiz_passing_score,
                order=sec.order or idx,
            )
            db.add(section)
            db.flush()

            if sec.documents:
                for doc in sec.documents:
                    db_doc = OnboardingDocument(section_id=section.id, **doc.model_dump())
                    db.add(db_doc)

            if sec.questions:
                for q in sec.questions:
                    db_q = OnboardingQuizQuestion(
                        section_id=section.id,
                        question=q.question,
                        options=q.options,
                        correct_option_index=q.correct_option_index
                    )
                    db.add(db_q)

    db.commit()
    db.refresh(module)
    return module


@router.delete("/modules/{module_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_module(
    module_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Delete a module. Restricted to Admins and PMs."""
    module = db.query(OnboardingModule).filter(OnboardingModule.id == module_id).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")
    db.delete(module)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


# ── Nested Sections CRUD ──────────────────────────────────────────

@router.post("/modules/{module_id}/sections", response_model=OnboardingSectionResponse, status_code=status.HTTP_201_CREATED)
def create_section(
    module_id: int,
    payload: OnboardingSectionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Create a section within a module, including nested documents and quiz questions."""
    module = db.query(OnboardingModule).filter(OnboardingModule.id == module_id).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")

    # Create section
    section = OnboardingSection(
        module_id=module_id,
        title=payload.title,
        description=payload.description,
        video_url=payload.video_url,
        video_duration=payload.video_duration,
        quiz_passing_score=payload.quiz_passing_score,
        order=payload.order,
    )
    db.add(section)
    db.flush()  # get section ID

    # Create nested documents
    if payload.documents:
        for doc in payload.documents:
            db_doc = OnboardingDocument(section_id=section.id, **doc.model_dump())
            db.add(db_doc)

    # Create nested quiz questions
    if payload.questions:
        for q in payload.questions:
            db_q = OnboardingQuizQuestion(
                section_id=section.id,
                question=q.question,
                options=q.options,
                correct_option_index=q.correct_option_index
            )
            db.add(db_q)

    db.commit()
    db.refresh(section)
    return section


@router.delete("/sections/{section_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_section(
    section_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Delete a section. Restricted to Admins and PMs."""
    section = db.query(OnboardingSection).filter(OnboardingSection.id == section_id).first()
    if not section:
        raise HTTPException(status_code=404, detail="Section not found")
    db.delete(section)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


# ── Excel Import / Templates ──────────────────────────────────────────

@router.post("/modules/{module_id}/sections/{section_id}/import-questions")
async def import_quiz_questions(
    module_id: int,
    section_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Import quiz questions from an Excel sheet into a specific section."""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=501, detail="Excel library (openpyxl) is not installed on the server.")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid Excel file.") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail="Empty sheet or missing headers.")

    headers = [str(h).strip().lower() for h in rows[0]]
    required = ['question', 'option1', 'option2', 'option3', 'option4', 'correctoption']
    for req in required:
        if req not in headers:
            raise HTTPException(status_code=400, detail=f"Missing required header column: {req}")

    q_idx = headers.index('question')
    o1_idx = headers.index('option1')
    o2_idx = headers.index('option2')
    o3_idx = headers.index('option3')
    o4_idx = headers.index('option4')
    c_idx = headers.index('correctoption')

    created = 0
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(v is None for v in row):
            continue

        question_text = str(row[q_idx]).strip() if row[q_idx] is not None else ""
        opt1 = str(row[o1_idx]).strip() if row[o1_idx] is not None else ""
        opt2 = str(row[o2_idx]).strip() if row[o2_idx] is not None else ""
        opt3 = str(row[o3_idx]).strip() if row[o3_idx] is not None else ""
        opt4 = str(row[o4_idx]).strip() if row[o4_idx] is not None else ""
        
        try:
            correct_opt = int(row[c_idx])
        except (ValueError, TypeError):
            errors.append(f"Row {row_num}: Invalid correctOption value.")
            continue

        if not question_text or not opt1 or not opt2 or not opt3 or not opt4 or correct_opt < 1 or correct_opt > 4:
            errors.append(f"Row {row_num}: Missing fields or correctOption not in range 1-4.")
            continue

        # Create question
        db_q = OnboardingQuizQuestion(
            section_id=section_id,
            question=question_text,
            options=[opt1, opt2, opt3, opt4],
            correct_option_index=correct_opt - 1
        )
        db.add(db_q)
        created += 1

    db.commit()
    return {"message": f"Successfully imported {created} questions", "created": created, "errors": errors}


# ── Progress Tracking ──────────────────────────────────────────────────

@router.post("/progress/section", response_model=OnboardingProgressResponse)
def record_progress(
    payload: OnboardingProgressCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Mark a section complete for an employee."""
    user_id = payload.user_id or current_user.id

    section = db.query(OnboardingSection).filter(OnboardingSection.id == payload.section_id).first()
    if not section:
        raise HTTPException(status_code=404, detail="Section not found")

    if current_user.role == "employee":
        if is_module_locked(user_id, section.module_id, db):
            raise HTTPException(status_code=403, detail="Cannot record progress for a locked module.")

        # If the section has quiz questions, the employee must pass the quiz first
        questions_count = db.query(OnboardingQuizQuestion).filter(OnboardingQuizQuestion.section_id == payload.section_id).count()
        if questions_count > 0:
            attempts = db.query(OnboardingQuizAttempt).filter(
                OnboardingQuizAttempt.user_id == user_id,
                OnboardingQuizAttempt.section_id == payload.section_id
            ).all()
            
            correct_count = sum(1 for a in attempts if a.is_correct)
            score = int((correct_count / questions_count) * 100) if questions_count > 0 else 0
            passing_score = section.quiz_passing_score or 0
            
            if score < passing_score:
                raise HTTPException(
                    status_code=400,
                    detail=f"Quiz passing score not met. You scored {score}%, but {passing_score}% is required."
                )

    existing = db.query(OnboardingProgress).filter(
        OnboardingProgress.user_id == user_id,
        OnboardingProgress.section_id == payload.section_id
    ).first()

    if existing:
        existing.completed_at = datetime.utcnow()
        db.commit()
        db.refresh(existing)
        return existing

    progress = OnboardingProgress(
        user_id=user_id,
        module_id=section.module_id,
        section_id=payload.section_id
    )
    db.add(progress)
    db.commit()
    db.refresh(progress)
    return progress


@router.get("/progress/{user_id}", response_model=List[OnboardingProgressResponse])
def get_user_progress(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Retrieve all completed sections for a candidate."""
    return db.query(OnboardingProgress).filter(OnboardingProgress.user_id == user_id).all()


# ── Quiz Submission ────────────────────────────────────────────────────

class QuizAnswerPayload(BaseModel):
    question_id: int
    chosen_index: int


class QuizSubmissionPayload(BaseModel):
    user_id: Optional[int] = None
    section_id: int
    answers: List[QuizAnswerPayload]


@router.post("/quiz/submit")
def submit_quiz(
    payload: QuizSubmissionPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Submit quiz answers, scores them, and updates attempts."""
    user_id = payload.user_id or current_user.id
    
    section = db.query(OnboardingSection).filter(OnboardingSection.id == payload.section_id).first()
    if not section:
        raise HTTPException(status_code=404, detail="Section not found")

    if current_user.role == "employee":
        if is_module_locked(user_id, section.module_id, db):
            raise HTTPException(status_code=403, detail="Cannot submit quiz for a locked module.")
    
    question_ids = [ans.question_id for ans in payload.answers]
    questions = db.query(OnboardingQuizQuestion).filter(OnboardingQuizQuestion.id.in_(question_ids)).all()
    question_map = {q.id: q for q in questions}

    correct_count = 0
    results = []

    for answer in payload.answers:
        question = question_map.get(answer.question_id)
        if not question:
            continue

        is_correct = question.correct_option_index == answer.chosen_index
        if is_correct:
            correct_count += 1

        # Record attempt
        existing_attempt = db.query(OnboardingQuizAttempt).filter(
            OnboardingQuizAttempt.user_id == user_id,
            OnboardingQuizAttempt.question_id == answer.question_id
        ).first()

        if existing_attempt:
            existing_attempt.chosen_index = answer.chosen_index
            existing_attempt.is_correct = is_correct
            existing_attempt.attempted_at = datetime.utcnow()
            results.append(existing_attempt)
        else:
            attempt = OnboardingQuizAttempt(
                user_id=user_id,
                section_id=payload.section_id,
                question_id=answer.question_id,
                chosen_index=answer.chosen_index,
                is_correct=is_correct
            )
            db.add(attempt)
            results.append(attempt)

    db.commit()
    score_percent = int((correct_count / len(payload.answers)) * 100) if payload.answers else 0

    return {
        "message": "Quiz submitted successfully",
        "score": score_percent,
        "correctCount": correct_count,
        "totalQuestions": len(payload.answers),
    }


# ── Team Members CRUD ─────────────────────────────────────────────────

@router.get("/team", response_model=List[OnboardingTeamMemberResponse])
def get_team_members(db: Session = Depends(get_db)):
    """Retrieve all team contacts sorted by department."""
    return db.query(OnboardingTeamMember).order_by(
        OnboardingTeamMember.department.asc(),
        OnboardingTeamMember.name.asc()
    ).all()


@router.post("/team", response_model=OnboardingTeamMemberResponse, status_code=status.HTTP_201_CREATED)
def create_team_member(
    payload: OnboardingTeamMemberCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Add a team member contact."""
    member = OnboardingTeamMember(**payload.model_dump())
    db.add(member)
    db.commit()
    db.refresh(member)
    return member


@router.put("/team/{member_id}", response_model=OnboardingTeamMemberResponse)
def update_team_member(
    member_id: int,
    payload: OnboardingTeamMemberCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Edit a team member contact."""
    member = db.query(OnboardingTeamMember).filter(OnboardingTeamMember.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Team member not found")

    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(member, key, value)

    db.commit()
    db.refresh(member)
    return member


@router.delete("/team/{member_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_team_member(
    member_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Remove a team member contact."""
    member = db.query(OnboardingTeamMember).filter(OnboardingTeamMember.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Team member not found")
    db.delete(member)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.get("/team/sample-excel")
def download_team_sample():
    """Download template Excel for team bulk import."""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=501, detail="Excel library (openpyxl) is not installed on the server.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Team Members"

    headers = ['name', 'role', 'department', 'email', 'linkedin', 'slack']
    ws.append(headers)

    ws.append(['Emily Rodriguez', 'HR Business Partner', 'Human Resources', 'emily@company.com', 'https://linkedin.com/in/emily', ''])
    ws.append(['Raj Malhotra', 'Engineering Manager', 'Engineering', 'raj@company.com', '', ''])

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=autonex_team_sample.xlsx"}
    )


@router.post("/team/bulk-import")
async def bulk_import_team(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Bulk import team members from an uploaded Excel file."""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=501, detail="Excel library (openpyxl) is not installed on the server.")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid Excel file.") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail="Empty sheet or missing headers.")

    headers = [str(h).strip().lower() for h in rows[0]]
    required = ['name', 'role', 'department']
    for req in required:
        if req not in headers:
            raise HTTPException(status_code=400, detail=f"Missing required header column: {req}")

    n_idx = headers.index('name')
    r_idx = headers.index('role')
    d_idx = headers.index('department')
    e_idx = headers.index('email') if 'email' in headers else None
    l_idx = headers.index('linkedin') if 'linkedin' in headers else None
    s_idx = headers.index('slack') if 'slack' in headers else None

    created = 0
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(v is None for v in row):
            continue

        name = str(row[n_idx]).strip() if row[n_idx] is not None else ""
        role = str(row[r_idx]).strip() if row[r_idx] is not None else ""
        dept = str(row[d_idx]).strip() if row[d_idx] is not None else ""
        email = str(row[e_idx]).strip() if e_idx is not None and row[e_idx] is not None else None
        linkedin = str(row[l_idx]).strip() if l_idx is not None and row[l_idx] is not None else None
        slack = str(row[s_idx]).strip() if s_idx is not None and row[s_idx] is not None else None

        if not name or not role or not dept:
            errors.append(f"Row {row_num}: Name, role, and department are required.")
            continue

        member = OnboardingTeamMember(
            name=name, role=role, department=dept,
            email=email, linkedin=linkedin, slack=slack
        )
        db.add(member)
        created += 1

    db.commit()
    return {"message": f"Bulk import complete: {created} contacts created", "created": created, "errors": errors}


# ── Candidates / Dashboard Endpoints ────────────────────────────────────

@router.get("/candidates/{user_id}/dashboard")
def get_candidate_dashboard(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Retrieve full onboarding progress metrics for a candidate."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Candidate not found")

    # Retrieve candidate employee details and mentor (PM)
    employee = db.query(Employee).filter(Employee.id == user.employee_id).first()
    mentor_name = "Unassigned"
    if employee and employee.converted_by:
        admin_pm = db.query(User).filter(User.id == employee.converted_by).first()
        if admin_pm:
            mentor_name = admin_pm.name

    # Fetch modules, progress, attempts
    modules = db.query(OnboardingModule).filter(OnboardingModule.status.ilike("PUBLISHED")).order_by(OnboardingModule.order.asc()).all()
    progress_records = db.query(OnboardingProgress).filter(OnboardingProgress.user_id == user_id).all()
    quiz_attempts = db.query(OnboardingQuizAttempt).filter(OnboardingQuizAttempt.user_id == user_id).all()

    completed_section_ids = {p.section_id for p in progress_records}
    completed_module_sections = {}
    for p in progress_records:
        completed_module_sections[p.module_id] = completed_module_sections.get(p.module_id, 0) + 1

    attempt_map = {a.question_id: a.is_correct for a in quiz_attempts}

    total_progress = 0
    total_questions = 0
    correct_answers = 0

    enriched_modules = []
    previous_completed = True
    for m in modules:
        total_sections = len(m.sections)
        completed = completed_module_sections.get(m.id, 0)
        progress = int((completed / total_sections) * 100) if total_sections > 0 else 0

        # Determine locked state
        is_locked = not previous_completed

        module_questions = 0
        module_correct = 0
        for s in m.sections:
            for q in s.questions:
                module_questions += 1
                total_questions += 1
                if attempt_map.get(q.id):
                    module_correct += 1
                    correct_answers += 1

        is_completed = (progress == 100 or total_sections == 0)
        status_str = "completed" if is_completed else "in_progress"
        total_progress += progress

        sections_list = []
        for s in m.sections:
            docs = [{"id": d.id, "title": d.title, "type": d.type, "url": d.url} for d in s.documents]
            qs = [{"id": q.id, "question": q.question, "options": q.options} for q in s.questions]
            sections_list.append({
                "id": s.id,
                "title": s.title,
                "description": s.description,
                "videoUrl": s.video_url,
                "videoDuration": s.video_duration,
                "quizPassingScore": s.quiz_passing_score,
                "order": s.order,
                "documents": docs,
                "questions": qs,
                "completed": s.id in completed_section_ids
            })

        enriched_modules.append({
            "id": m.id,
            "title": m.title,
            "description": m.description,
            "status": status_str,
            "progress": progress,
            "totalLessons": total_sections,
            "completedLessons": completed,
            "assessmentUrl": m.assessment_url,
            "sections": sections_list,
            "quizScore": int((module_correct / module_questions) * 100) if module_questions > 0 else 0,
            "locked": is_locked
        })

        # Update previous_completed state for the next module
        previous_completed = is_completed

    overall_progress = int(total_progress / len(modules)) if modules else 0
    avg_quiz_score = int((correct_answers / total_questions) * 100) if total_questions > 0 else 0

    return {
        "user": {
            "id": user.id,
            "name": user.name,
            "email": user.email,
            "department": employee.employee_type if employee else "Engineering"
        },
        "modules": enriched_modules,
        "stats": {
            "overallProgress": overall_progress,
            "completedModules": len([m for m in enriched_modules if m["status"] == "completed"]),
            "totalModules": len(modules),
            "avgQuizScore": avg_quiz_score,
            "mentorName": mentor_name
        }
    }


# ── Analytics / KPI Dashboard ──────────────────────────────────────────

@router.get("/analytics/dashboard")
def get_analytics_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Retrieve high-level onboarding KPIs and recent registrations."""
    candidates = get_onboarding_candidates(db)
    modules_count = db.query(OnboardingModule).filter(OnboardingModule.status.ilike("PUBLISHED")).count()

    recent = []
    for c in candidates[:5]:
        employee = db.query(Employee).filter(Employee.id == c.employee_id).first()
        recent.append({
            "id": c.id,
            "name": c.name,
            "email": c.email,
            "department": employee.designation if employee else "Annotator",
            "createdAt": c.created_at
        })

    # Average progress computation
    total_progress = 0
    for c in candidates:
        completed_sections = db.query(OnboardingProgress).filter(OnboardingProgress.user_id == c.id).count()
        total_sections = db.query(OnboardingSection).join(OnboardingModule).filter(
            OnboardingModule.status.ilike("PUBLISHED")
        ).count()
        progress = int((completed_sections / total_sections) * 100) if total_sections > 0 else 0
        total_progress += progress

    avg_progress = int(total_progress / len(candidates)) if candidates else 0

    return {
        "metrics": {
            "totalCandidates": len(candidates),
            "avgProgress": avg_progress,
            "modulesCompleted": len([c for c in candidates if db.query(OnboardingProgress).filter(OnboardingProgress.user_id == c.id).count() > 0]),
            "totalModules": modules_count
        },
        "recentCandidates": recent
    }


@router.get("/analytics/full")
def get_full_analytics(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Retrieve full analytics data distributions and weekly metrics."""
    candidates = get_onboarding_candidates(db)
    modules = db.query(OnboardingModule).filter(OnboardingModule.status.ilike("PUBLISHED")).all()

    total_progress_sum = 0
    total_score_sum = 0
    distribution = {"0-25%": 0, "26-50%": 0, "51-75%": 0, "76-100%": 0}

    for c in candidates:
        completed_sections = db.query(OnboardingProgress).filter(OnboardingProgress.user_id == c.id).count()
        total_sections = db.query(OnboardingSection).join(OnboardingModule).filter(
            OnboardingModule.status.ilike("PUBLISHED")
        ).count()
        progress = int((completed_sections / total_sections) * 100) if total_sections > 0 else 0
        total_progress_sum += progress

        if progress <= 25:
            distribution["0-25%"] += 1
        elif progress <= 50:
            distribution["26-50%"] += 1
        elif progress <= 75:
            distribution["51-75%"] += 1
        else:
            distribution["76-100%"] += 1

        # Quiz calculations
        attempts = db.query(OnboardingQuizAttempt).filter(OnboardingQuizAttempt.user_id == c.id).all()
        correct = len([a for a in attempts if a.is_correct])
        score = int((correct / len(attempts)) * 100) if attempts else 0
        total_score_sum += score

    num_candidates = len(candidates) or 1
    avg_score = int(total_score_sum / num_candidates)
    completion_rate = int(total_progress_sum / num_candidates)

    return {
        "kpis": [
            {"label": "Avg Quiz Score", "value": f"{avg_score}%", "trend": "+2.4%"},
            {"label": "Completion Rate", "value": f"{completion_rate}%", "trend": "+5.1%"},
            {"label": "Avg Time to Complete", "value": "14 Days", "trend": "-1.2 Days"}
        ],
        "weeklyData": [
            {"name": "Mon", "completion": 2},
            {"name": "Tue", "completion": 3},
            {"name": "Wed", "completion": 1},
            {"name": "Thu", "completion": 5},
            {"name": "Fri", "completion": 4},
            {"name": "Sat", "completion": 0},
            {"name": "Sun", "completion": 1}
        ],
        "distribution": [
            {"name": "0-25%", "value": distribution["0-25%"]},
            {"name": "26-50%", "value": distribution["26-50%"]},
            {"name": "51-75%", "value": distribution["51-75%"]},
            {"name": "76-100%", "value": distribution["76-100%"]}
        ]
    }


# ── Audit Progress Reports Endpoints ───────────────────────────────────

def fetch_onboarding_reports_data(db: Session) -> List[dict]:
    candidates = get_onboarding_candidates(db)
    modules = db.query(OnboardingModule).filter(OnboardingModule.status.ilike("PUBLISHED")).all()

    reports = []
    for c in candidates:
        employee = db.query(Employee).filter(Employee.id == c.employee_id).first()
        dept = employee.designation if employee else "Annotator"

        progress_records = db.query(OnboardingProgress).filter(OnboardingProgress.user_id == c.id).all()
        completed_section_ids = {p.section_id for p in progress_records}
        completed_module_sections = {}
        for p in progress_records:
            completed_module_sections[p.module_id] = completed_module_sections.get(p.module_id, 0) + 1

        quiz_attempts = db.query(OnboardingQuizAttempt).filter(OnboardingQuizAttempt.user_id == c.id).all()
        attempt_map = {a.question_id: a.is_correct for a in quiz_attempts}

        total_progress = 0
        total_questions = 0
        correct_answers = 0
        attempted_questions = len(quiz_attempts)

        module_stats = []
        for m in modules:
            total_sections = len(m.sections)
            completed = completed_module_sections.get(m.id, 0)
            progress = int((completed / total_sections) * 100) if total_sections > 0 else 0
            total_progress += progress

            module_questions = 0
            module_correct = 0
            for s in m.sections:
                for q in s.questions:
                    module_questions += 1
                    total_questions += 1
                    if attempt_map.get(q.id):
                        module_correct += 1
                        correct_answers += 1

            score = int((module_correct / module_questions) * 100) if module_questions > 0 else 0
            module_stats.append({
                "moduleId": m.id,
                "moduleTitle": m.title,
                "progress": progress,
                "score": score,
                "attemptedQuestions": module_questions,
                "totalQuestions": module_questions,
                "correctAnswers": module_correct
            })

        overall_progress = int(total_progress / len(modules)) if modules else 0
        overall_score = int((correct_answers / total_questions) * 100) if total_questions > 0 else 0

        reports.append({
            "userId": c.id,
            "name": c.name,
            "email": c.email,
            "department": dept,
            "overallProgress": overall_progress,
            "overallScore": overall_score,
            "attemptedQuestions": attempted_questions,
            "correctAnswers": correct_answers,
            "totalQuestions": total_questions,
            "moduleStats": module_stats
        })

    return reports


@router.get("/reports")
def get_onboarding_reports(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Retrieve full detailed reports on candidate onboarding progress."""
    return fetch_onboarding_reports_data(db)


@router.get("/reports/export")
def export_onboarding_reports(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Export onboarding progress report as a CSV file download."""
    reports = fetch_onboarding_reports_data(db)

    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_NONNUMERIC)

    # Headers
    writer.writerow([
        'Candidate Name', 'Email', 'Department', 'Overall Progress %', 'Overall Quiz Score %', 
        'Correct Answers', 'Total Questions', 'Module', 'Module Progress %', 'Module Score %', 
        'Module Correct', 'Module Total Qs'
    ])

    for r in reports:
        if r["moduleStats"]:
            for ms in r["moduleStats"]:
                writer.writerow([
                    r["name"], r["email"], r["department"], r["overallProgress"], r["overallScore"], 
                    r["correctAnswers"], r["totalQuestions"], ms["moduleTitle"], ms["progress"], 
                    ms["score"], ms["correctAnswers"], ms["totalQuestions"]
                ])
        else:
            writer.writerow([
                r["name"], r["email"], r["department"], r["overallProgress"], r["overallScore"], 
                r["correctAnswers"], r["totalQuestions"], "", "", "", "", ""
            ])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=autonex_candidates_report.csv"}
    )


@router.get("/mentors/{mentor_id}/mentees")
def get_mentees(
    mentor_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "pm"))
):
    """Retrieve all candidate employees assigned to a PM (mentor) as mentees or allocated to their annotation projects."""
    pm_user = db.query(User).filter(User.id == mentor_id).first()
    if not pm_user or not pm_user.employee_id:
        return []

    pm_employee_id = pm_user.employee_id

    # 1. Fetch direct mentees (where Employee.mentor_id == pm_employee_id)
    mentees = db.query(User).join(
        Employee, User.employee_id == Employee.id
    ).filter(
        Employee.mentor_id == pm_employee_id
    ).all()

    # 2. Fetch team members allocated to annotation sub-projects managed by this PM
    from app.models.parent_project import MainProject
    from app.models.project import DailySheet
    from app.models.allocation import Allocation

    # Find all daily sheets (sub-projects in UI) where is_annotation == True
    annotation_sheets = db.query(DailySheet).filter(DailySheet.is_annotation == True).all()
    
    pm_sheet_ids = []
    for sheet in annotation_sheets:
        if not sheet.main_project_id:
            continue
        
        # Check if the PM is the program manager of the parent project
        parent_project = db.query(MainProject).filter(MainProject.id == sheet.main_project_id).first()
        if parent_project:
            pms = parent_project.program_manager_ids or []
            if not pms and parent_project.program_manager_id:
                pms = [parent_project.program_manager_id]
            if pm_employee_id in pms:
                pm_sheet_ids.append(sheet.id)

    allocated_employee_ids = set()
    if pm_sheet_ids:
        # Find allocations for these daily sheets
        allocs = db.query(Allocation).filter(Allocation.sub_project_id.in_(pm_sheet_ids)).all()
        allocated_employee_ids = {a.employee_id for a in allocs}

    allocated_users = []
    if allocated_employee_ids:
        allocated_users = db.query(User).join(
            Employee, User.employee_id == Employee.id
        ).filter(
            (User.employee_id.in_(list(allocated_employee_ids))) &
            ((Employee.designation.ilike("%Annotator%")) | (Employee.designation.ilike("%Reviewer%")))
        ).all()

    # Union the direct mentees and the allocated annotators/reviewers
    user_map = {u.id: u for u in mentees}
    for u in allocated_users:
        user_map[u.id] = u
    final_users = list(user_map.values())

    results = []
    for m in final_users:
        progress_records = db.query(OnboardingProgress).filter(OnboardingProgress.user_id == m.id).all()
        completed_modules_count = len({p.module_id for p in progress_records})

        attempts = db.query(OnboardingQuizAttempt).filter(OnboardingQuizAttempt.user_id == m.id).all()
        correct = len([a for a in attempts if a.is_correct])
        score = int((correct / len(attempts)) * 100) if attempts else 0

        employee = db.query(Employee).filter(Employee.id == m.employee_id).first()

        results.append({
            "id": m.id,
            "name": m.name,
            "email": m.email,
            "department": employee.designation if employee else "Annotator",
            "isActive": m.is_active,
            "completedModulesCount": completed_modules_count,
            "quizScorePercent": score
        })

    return results

